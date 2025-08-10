from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db.models import Count, Sum, Q, Avg
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.template.loader import render_to_string
from datetime import datetime, timedelta
import json
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import base64

from .models import (
    UserProfile, Department, Service, Appointment, 
    TestResult, Payment, MedicalCertificate, Notification, 
    AuditLog, SystemSettings, ChatbotConversation
)
from firebase_config import firebase_config

def create_audit_log(request, action, model_name, object_id='', changes=None):
    """Create audit log entry"""
    try:
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=action,
            model_name=model_name,
            object_id=str(object_id),
            changes=changes or {},
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            session_key=request.session.session_key
        )
    except Exception as e:
        print(f"Error creating audit log: {e}")

def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def home(request):
    """Enhanced homepage with comprehensive statistics and modern design"""
    try:
        # Get comprehensive statistics
        total_patients = User.objects.filter(userprofile__role='patient').count()
        total_appointments = Appointment.objects.count()
        total_services = Service.objects.filter(is_available=True).count()
        total_departments = Department.objects.filter(is_active=True).count()
        
        # Today's statistics
        today = timezone.now().date()
        today_appointments = Appointment.objects.filter(appointment_date__date=today).count()
        pending_appointments = Appointment.objects.filter(status='pending').count()
        
        # Revenue statistics
        total_revenue = Payment.objects.filter(payment_status='completed').aggregate(
            total=Sum('amount'))['total'] or 0
        
        # Recent data
        departments = Department.objects.filter(is_active=True).prefetch_related('services')[:6]
        services = Service.objects.filter(is_available=True).select_related('department')[:8]
        recent_appointments = Appointment.objects.select_related(
            'patient', 'service'
        ).order_by('-created_at')[:5]
        
        # Patient satisfaction (mock data - can be replaced with real survey data)
        satisfaction_rate = 98.5
        
        # Monthly appointment trends for chart
        monthly_data = []
        for i in range(6):
            month_start = (today.replace(day=1) - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            count = Appointment.objects.filter(
                appointment_date__date__range=[month_start, month_end]
            ).count()
            monthly_data.append({
                'month': month_start.strftime('%b %Y'),
                'count': count
            })
        monthly_data.reverse()
        
        # Service popularity
        popular_services = Service.objects.annotate(
            booking_count=Count('appointments')
        ).order_by('-booking_count')[:5]
        
        context = {
            'stats': {
                'total_patients': total_patients,
                'total_appointments': total_appointments,
                'total_services': total_services,
                'total_departments': total_departments,
                'today_appointments': today_appointments,
                'pending_appointments': pending_appointments,
                'total_revenue': total_revenue,
                'satisfaction_rate': satisfaction_rate,
            },
            'departments': departments,
            'services': services,
            'recent_appointments': recent_appointments,
            'monthly_data': monthly_data,
            'popular_services': popular_services,
            'firebase_config': settings.FIREBASE_CONFIG,
        }
        
        # Create audit log
        create_audit_log(request, 'view', 'Homepage')
        
        return render(request, 'hospital/home.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading homepage: {str(e)}')
        return render(request, 'hospital/home.html', {'stats': {}, 'departments': [], 'services': []})

def login_view(request):
    """Enhanced login with Google OAuth integration"""
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')
        
        if not username or not password:
            messages.error(request, 'Please enter both username and password.')
            return render(request, 'hospital/login.html')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                
                # Set session expiry based on remember me
                if not remember_me:
                    request.session.set_expiry(0)  # Browser session
                else:
                    request.session.set_expiry(1209600)  # 2 weeks
                
                # Create audit log
                create_audit_log(request, 'login', 'User', user.id)
                
                # Redirect based on user role
                try:
                    profile = user.userprofile
                    if profile.role == 'admin':
                        messages.success(request, f'Welcome back, {user.get_full_name()}!')
                        return redirect('admin_dashboard')
                    elif profile.role == 'patient':
                        messages.success(request, f'Welcome back, {user.get_full_name()}!')
                        return redirect('patient_dashboard')
                    else:
                        return redirect('home')
                except UserProfile.DoesNotExist:
                    messages.warning(request, 'Profile not found. Please contact administrator.')
                    return redirect('home')
            else:
                messages.error(request, 'Your account has been deactivated. Please contact administrator.')
        else:
            messages.error(request, 'Invalid username or password.')
            create_audit_log(request, 'login', 'User', username, {'status': 'failed'})
    
    context = {
        'firebase_config': settings.FIREBASE_CONFIG,
        'google_client_id': settings.GOOGLE_OAUTH2_CLIENT_ID,
    }
    
    return render(request, 'hospital/login.html', context)

def register_view(request):
    """Enhanced registration with validation and Firebase integration"""
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        # Get form data
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone_number = request.POST.get('phone_number', '').strip()
        
        # Validation
        errors = []
        
        if not all([username, email, password, confirm_password, first_name, last_name]):
            errors.append('All fields are required.')
        
        if password != confirm_password:
            errors.append('Passwords do not match.')
        
        if len(password) &lt; 8:
            errors.append('Password must be at least 8 characters long.')
        
        if User.objects.filter(username=username).exists():
            errors.append('Username already exists.')
        
        if User.objects.filter(email=email).exists():
            errors.append('Email already registered.')
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'hospital/register.html')
        
        try:
            # Create user
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            
            # Update user profile
            profile = user.userprofile
            profile.role = 'patient'
            profile.phone_number = phone_number
            profile.save()
            
            # Create Firebase user if available
            try:
                firebase_user = firebase_config.create_user(
                    email=email,
                    password=password,
                    display_name=f"{first_name} {last_name}"
                )
                if firebase_user:
                    profile.firebase_uid = firebase_user.uid
                    profile.save()
            except Exception as e:
                print(f"Firebase user creation failed: {e}")
            
            # Create audit log
            create_audit_log(request, 'create', 'User', user.id)
            
            # Send welcome email
            if not settings.DEBUG:
                try:
                    send_mail(
                        'Welcome to MAES Laboratory',
                        f'Dear {first_name},\n\nWelcome to MAES Laboratory Management System. Your account has been created successfully.',
                        settings.DEFAULT_FROM_EMAIL,
                        [email],
                        fail_silently=True,
                    )
                except:
                    pass
            
            messages.success(request, 'Registration successful! Please log in with your credentials.')
            return redirect('login')
            
        except Exception as e:
            messages.error(request, 'Registration failed. Please try again.')
            return render(request, 'hospital/register.html')
    
    context = {
        'firebase_config': settings.FIREBASE_CONFIG,
    }
    
    return render(request, 'hospital/register.html', context)

@login_required
def patient_dashboard(request):
    """Enhanced patient dashboard with comprehensive information and charts"""
    if request.user.userprofile.role != 'patient':
        messages.error(request, 'Access denied. Patient access required.')
        return redirect('home')
    
    # Get user appointments
    appointments = Appointment.objects.filter(patient=request.user).order_by('-appointment_date')
    
    # Pagination
    paginator = Paginator(appointments, 10)
    page_number = request.GET.get('page')
    appointments_page = paginator.get_page(page_number)
    
    # Get test results
    recent_results = TestResult.objects.filter(
        appointment__patient=request.user,
        status='released'
    ).order_by('-released_at')[:5]
    
    # Get notifications
    notifications = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).order_by('-created_at')[:5]
    
    # Get statistics
    stats = {
        'total_appointments': appointments.count(),
        'pending_appointments': appointments.filter(status='pending').count(),
        'completed_appointments': appointments.filter(status='completed').count(),
        'upcoming_appointments': appointments.filter(
            appointment_date__gte=timezone.now(),
            status__in=['pending', 'confirmed']
        ).count(),
    }
    
    # Get upcoming appointment
    next_appointment = appointments.filter(
        appointment_date__gte=timezone.now(),
        status__in=['pending', 'confirmed']
    ).first()
    
    # Generate appointment history chart
    appointment_chart = generate_patient_appointment_chart(request.user)
    
    # Get payment history
    payments = Payment.objects.filter(
        appointment__patient=request.user
    ).order_by('-payment_date')[:5]
    
    context = {
        'appointments': appointments_page,
        'recent_results': recent_results,
        'notifications': notifications,
        'stats': stats,
        'next_appointment': next_appointment,
        'appointment_chart': appointment_chart,
        'payments': payments,
    }
    
    # Create audit log
    create_audit_log(request, 'view', 'PatientDashboard')
    
    return render(request, 'hospital/patient_dashboard.html', context)

@login_required
def admin_dashboard(request):
    """Enhanced admin dashboard with comprehensive analytics and visualizations"""
    if request.user.userprofile.role != 'admin':
        messages.error(request, 'Access denied. Administrator access required.')
        return redirect('home')
    
    # Date ranges
    today = timezone.now().date()
    this_week = today - timedelta(days=7)
    this_month = today.replace(day=1)
    last_month = (this_month - timedelta(days=1)).replace(day=1)
    
    # Comprehensive statistics
    stats = {
        'total_patients': User.objects.filter(userprofile__role='patient').count(),
        'total_appointments': Appointment.objects.count(),
        'today_appointments': Appointment.objects.filter(appointment_date__date=today).count(),
        'pending_appointments': Appointment.objects.filter(status='pending').count(),
        'completed_appointments': Appointment.objects.filter(status='completed').count(),
        'total_revenue': Payment.objects.filter(payment_status='completed').aggregate(
            total=Sum('amount'))['total'] or 0,
        'monthly_revenue': Payment.objects.filter(
            payment_date__gte=this_month,
            payment_status='completed'
        ).aggregate(total=Sum('amount'))['total'] or 0,
        'weekly_appointments': Appointment.objects.filter(
            created_at__gte=this_week
        ).count(),
        'total_services': Service.objects.filter(is_available=True).count(),
        'total_departments': Department.objects.filter(is_active=True).count(),
    }
    
    # Recent activities
    recent_appointments = Appointment.objects.select_related(
        'patient', 'service'
    ).order_by('-created_at')[:10]
    
    recent_payments = Payment.objects.select_related(
        'appointment__patient'
    ).order_by('-payment_date')[:10]
    
    recent_results = TestResult.objects.select_related(
        'appointment__patient', 'appointment__service'
    ).order_by('-created_at')[:5]
    
    # Department statistics
    department_stats = Department.objects.annotate(
        appointment_count=Count('services__appointments'),
        revenue=Sum('services__appointments__payments__amount')
    ).filter(is_active=True)
    
    # Generate charts
    charts = {
        'monthly_appointments': generate_monthly_appointments_chart(),
        'revenue_chart': generate_revenue_chart(),
        'service_popularity': generate_service_popularity_chart(),
        'patient_demographics': generate_patient_demographics_chart(),
    }
    
    # Service popularity
    popular_services = Service.objects.annotate(
        booking_count=Count('appointments')
    ).order_by('-booking_count')[:5]
    
    # Recent registrations
    recent_patients = User.objects.filter(
        userprofile__role='patient'
    ).order_by('-date_joined')[:5]
    
    context = {
        'stats': stats,
        'recent_appointments': recent_appointments,
        'recent_payments': recent_payments,
        'recent_results': recent_results,
        'department_stats': department_stats,
        'popular_services': popular_services,
        'recent_patients': recent_patients,
        'charts': charts,
    }
    
    # Create audit log
    create_audit_log(request, 'view', 'AdminDashboard')
    
    return render(request, 'hospital/admin_dashboard.html', context)

@login_required
def book_appointment(request):
    """Enhanced appointment booking with financial assistance and validation"""
    if request.method == 'POST':
        service_id = request.POST.get('service')
        appointment_date = request.POST.get('appointment_date')
        appointment_time = request.POST.get('appointment_time')
        notes = request.POST.get('notes', '')
        financial_assistance = request.POST.get('financial_assistance', 'none')
        insurance_details = request.POST.get('insurance_details', '')
        
        # Validation
        if not all([service_id, appointment_date, appointment_time]):
            messages.error(request, 'Please fill in all required fields.')
            return redirect('book_appointment')
        
        try:
            service = get_object_or_404(Service, id=service_id, is_available=True)
            
            # Combine date and time
            appointment_datetime = datetime.strptime(f"{appointment_date} {appointment_time}", "%Y-%m-%d %H:%M")
            appointment_datetime = timezone.make_aware(appointment_datetime)
            
            # Check if appointment is in the future
            if appointment_datetime &lt;= timezone.now():
                messages.error(request, 'Please select a future date and time.')
                return redirect('book_appointment')
            
            # Check for conflicts
            existing_appointment = Appointment.objects.filter(
                patient=request.user,
                appointment_date=appointment_datetime,
                status__in=['pending', 'confirmed']
            ).exists()
            
            if existing_appointment:
                messages.error(request, 'You already have an appointment at this time.')
                return redirect('book_appointment')
            
            # Calculate pricing
            total_amount = service.price
            discount_amount = 0
            
            # Apply financial assistance discount
            if financial_assistance == 'hmo':
                discount_amount = total_amount * 0.8  # 80% HMO coverage
            elif financial_assistance == 'senior':
                discount_amount = total_amount * 0.2  # 20% senior discount
            elif financial_assistance == 'pwd':
                discount_amount = total_amount * 0.2  # 20% PWD discount
            elif financial_assistance == 'student':
                discount_amount = total_amount * 0.15  # 15% student discount
            
            # Create appointment
            appointment = Appointment.objects.create(
                patient=request.user,
                service=service,
                appointment_date=appointment_datetime,
                total_amount=total_amount,
                discount_amount=discount_amount,
                final_amount=total_amount - discount_amount,
                notes=notes,
                financial_assistance_type=financial_assistance,
                insurance_details={'details': insurance_details} if insurance_details else {}
            )
            
            # Create notification
            Notification.objects.create(
                user=request.user,
                notification_type='appointment_confirmed',
                title='Appointment Booked Successfully',
                message=f'Your appointment for {service.name} on {appointment_datetime.strftime("%B %d, %Y at %I:%M %p")} has been booked.',
                related_appointment=appointment
            )
            
            # Create audit log
            create_audit_log(request, 'create', 'Appointment', appointment.appointment_id)
            
            messages.success(request, f'Appointment booked successfully! Your appointment ID is {appointment.appointment_id}.')
            return redirect('patient_dashboard')
            
        except Exception as e:
            messages.error(request, f'Failed to book appointment: {str(e)}')
            return redirect('book_appointment')
    
    # GET request - show booking form
    services = Service.objects.filter(is_available=True).select_related('department')
    departments = Department.objects.filter(is_active=True)
    
    context = {
        'services': services,
        'departments': departments,
    }
    
    return render(request, 'hospital/book_appointment.html', context)

@login_required
def medical_certificate_request(request):
    """Request medical certificate"""
    if request.method == 'POST':
        certificate_type = request.POST.get('certificate_type')
        purpose = request.POST.get('purpose')
        valid_from = request.POST.get('valid_from')
        valid_until = request.POST.get('valid_until')
        
        try:
            # Create medical certificate request
            certificate = MedicalCertificate.objects.create(
                patient=request.user,
                certificate_type=certificate_type,
                purpose=purpose,
                valid_from=valid_from,
                valid_until=valid_until if valid_until else None,
                issued_by=request.user,  # Will be changed by admin
            )
            
            # Create notification for admin
            admin_users = User.objects.filter(userprofile__role='admin')
            for admin in admin_users:
                Notification.objects.create(
                    user=admin,
                    notification_type='other',
                    title='New Medical Certificate Request',
                    message=f'{request.user.get_full_name()} has requested a {certificate.get_certificate_type_display()}.',
                    related_certificate=certificate
                )
            
            messages.success(request, 'Medical certificate request submitted successfully!')
            return redirect('patient_dashboard')
            
        except Exception as e:
            messages.error(request, f'Failed to submit request: {str(e)}')
    
    return render(request, 'hospital/medical_certificate_request.html')

@csrf_exempt
@require_http_methods(["POST"])
def chatbot_api(request):
    """Enhanced chatbot API with comprehensive responses and live admin support"""
    try:
        data = json.loads(request.body)
        message = data.get('message', '').lower().strip()
        session_id = data.get('session_id', request.session.session_key or 'anonymous')
        
        if not message:
            return JsonResponse({'response': 'Please type a message.'})
        
        # Enhanced responses with more comprehensive coverage
        responses = {
            # Greetings and basic interactions
            'hello': 'Hello! Welcome to MAES Laboratory. I\'m your virtual assistant. How can I help you today?',
            'hi': 'Hi there! I\'m here to help you with any questions about our laboratory services, appointments, or general inquiries.',
            'good morning': 'Good morning! How can I assist you with your healthcare needs today?',
            'good afternoon': 'Good afternoon! What can I do for you?',
            'good evening': 'Good evening! How may I assist you?',
            'help': 'I can help you with:\n• Booking appointments\n• Service information\n• Payment options\n• Test results\n• Operating hours\n• Contact information\n\nWhat would you like to know?',
            
            # Services and tests
            'services': 'We offer comprehensive laboratory services including:\n• Blood Tests (CBC, Lipid Profile, Blood Sugar)\n• Imaging (X-ray, Ultrasound, CT Scan)\n• Cardiac Tests (ECG, Stress Test)\n• Microscopy and Pathology\n• Drug Testing\n• Vaccination Services\n• DNA Testing\n\nWould you like details about any specific service?',
            'blood test': 'Our blood tests include:\n• Complete Blood Count (CBC) - ₱500\n• Lipid Profile - ₱800\n• Blood Sugar (FBS/RBS) - ₱300\n• Liver Function Test - ₱1,200\n• Kidney Function Test - ₱900\n• Thyroid Function Test - ₱1,500\n\nMost require 8-12 hours fasting. Results available in 24-48 hours.',
            'xray': 'Digital X-ray services available:\n• Chest X-ray - ₱800\n• Bone X-ray - ₱600\n• Spine X-ray - ₱1,000\n• Duration: 15 minutes\n• No special preparation required\n• Results available immediately',
            'ultrasound': 'Ultrasound imaging services:\n• Abdominal Ultrasound - ₱1,200\n• Pelvic Ultrasound - ₱1,000\n• Cardiac Echo - ₱2,500\n• Duration: 30-45 minutes\n• Some require fasting or full bladder',
            'ecg': 'Electrocardiogram (ECG) testing:\n• Resting ECG - ₱600\n• Stress Test ECG - ₱2,000\n• Duration: 20 minutes\n• No special preparation\n• Results available immediately',
            'ct scan': 'CT Scan services:\n• Head CT - ₱3,500\n• Chest CT - ₱4,000\n• Abdominal CT - ₱4,500\n• With contrast available\n• Appointment required',
            
            # Appointments and booking
            'appointment': 'To book an appointment:\n1. Visit our website and login\n2. Select "Book Appointment"\n3. Choose your service and preferred time\n4. Confirm your booking\n\nOr call us at (043) 286-2531\nOnline booking available 24/7!',
            'book': 'You can book appointments:\n• Online through our website (24/7)\n• Call (043) 286-2531\n• Walk-in (subject to availability)\n\nWe recommend booking in advance for guaranteed slots.',
            'schedule': 'Our operating hours:\n• Monday to Saturday: 8:00 AM - 6:00 PM\n• Sunday: CLOSED\n• Holidays: CLOSED\n\nEmergency services available through partner hospitals.',
            'cancel': 'To cancel an appointment:\n• Login to your account and go to dashboard\n• Call us at (043) 286-2531\n• At least 24 hours notice preferred\n• Cancellation fees may apply for same-day cancellations',
            
            # Location and contact
            'location': 'MAES Laboratory is located in Batangas City, Philippines.\n\nFor exact address and directions:\n• Check our website contact page\n• Call (043) 286-2531\n• We provide detailed directions and landmarks',
            'address': 'We\'re located in Batangas City. For the complete address and GPS coordinates, please visit our contact page or call (043) 286-2531.',
            'hours': 'Operating Hours:\n• Monday - Saturday: 8:00 AM to 6:00 PM\n• Sunday: CLOSED\n• Holidays: CLOSED\n\nAppointment booking available online 24/7',
            'contact': 'Contact Information:\n• Phone: (043) 286-2531\n• Email: info@maeslaboratory.com\n• Website: www.maeslaboratory.com\n• Available: Mon-Sat, 8AM-6PM',
            
            # Payment and pricing
            'payment': 'We accept multiple payment methods:\n• Cash\n• GCash and PayMaya\n• Bank Transfer\n• Credit/Debit Cards\n• HMO and Insurance\n• Installment plans available\n• Senior/PWD discounts: 20% off',
            'price': 'Our competitive pricing:\n• Blood tests: ₱300-₱1,500\n• X-rays: ₱600-₱1,000\n• Ultrasound: ₱1,000-₱2,500\n• ECG: ₱600-₱2,000\n• CT Scan: ₱3,500-₱4,500\n\nDiscounts available for seniors, PWDs, and HMO members.',
            'cost': 'Service costs vary by test type. We offer:\n• Competitive pricing\n• Package deals\n• HMO coverage\n• Senior/PWD discounts\n• Flexible payment options\n\nCall (043) 286-2531 for specific pricing.',
            'insurance': 'Insurance and HMO accepted:\n• PhilHealth\n• Major HMO providers\n• Corporate accounts\n• Bring valid ID and insurance card\n• Pre-authorization may be required',
            'hmo': 'HMO services:\n• Most major HMOs accepted\n• 80% coverage on most tests\n• Bring HMO card and valid ID\n• Some tests require pre-authorization\n• Check with your HMO for coverage details',
            
            # Results and reports
            'results': 'Test results:\n• Available within 24-48 hours\n• SMS/Email notification when ready\n• Online portal access\n• Physical copy pickup available\n• Rush results available for urgent cases',
            'report': 'Laboratory reports:\n• Digital copies via patient portal\n• Physical copies at our facility\n• Secure and confidential\n• Doctor interpretation available\n• Historical results accessible online',
            'when': 'Result turnaround times:\n• Blood tests: 24-48 hours\n• X-rays: Immediate\n• Ultrasound: Same day\n• CT Scan: 24 hours\n• Complex tests: 3-5 days\n\nRush processing available for urgent cases.',
            
            # Preparation and requirements
            'fasting': 'Fasting requirements:\n• Blood sugar tests: 8-12 hours\n• Lipid profile: 12 hours\n• Liver function: 8 hours\n• Water is allowed\n• Medications as prescribed by doctor\n• We\'ll inform you of specific requirements when booking',
            'preparation': 'Test preparation varies:\n• Some require fasting\n• Others need full bladder\n• Medication adjustments may be needed\n• Detailed instructions provided when booking\n• Call if you have questions about preparation',
            'requirements': 'What to bring:\n• Valid government ID\n• Doctor\'s request (if any)\n• HMO card (if applicable)\n• Previous test results (for comparison)\n• List of current medications',
            
            # Emergency and urgent care
            'emergency': 'For medical emergencies:\n• Call 911 immediately\n• Go to nearest emergency room\n• We provide diagnostic services, not emergency care\n• Partner hospitals available for urgent cases',
            'urgent': 'For urgent test results:\n• Rush processing available\n• Same-day results for most tests\n• Additional fees may apply\n• Call (043) 286-2531 for urgent requests',
            
            # Technology and quality
            'technology': 'Our advanced technology:\n• AI-powered analysis systems\n• Digital imaging equipment\n• Automated laboratory systems\n• 99.9% accuracy guarantee\n• ISO certified processes',
            'quality': 'Quality assurance:\n• ISO 15189 certified\n• Regular equipment calibration\n• Experienced technicians\n• Doctor-reviewed results\n• Continuous quality monitoring',
            'accuracy': 'We guarantee 99.9% accuracy through:\n• State-of-the-art equipment\n• Rigorous quality control\n• Expert technicians\n• Double-checking procedures\n• Regular proficiency testing',
            
            # COVID-19 related
            'covid': 'COVID-19 services:\n• RT-PCR testing available\n• Antigen rapid tests\n• Antibody testing\n• Safety protocols in place\n• Results within 24 hours for PCR',
            'safety': 'Safety measures:\n• Mandatory masks\n• Temperature screening\n• Social distancing\n• Regular disinfection\n• Limited capacity\n• Online booking preferred',
            
            # Special requests
            'live admin': 'I\'ll connect you with a live administrator. Please wait while I transfer your request...',
            'human': 'Let me connect you with our human support team. They\'ll be with you shortly.',
            'complaint': 'I\'m sorry to hear about your concern. Let me connect you with our customer service team who can address your complaint properly.',
        }
        
        # Find matching response
        response = None
        suggestions = []
        
        # Check for exact matches first
        for keyword, reply in responses.items():
            if keyword in message:
                response = reply
                break
        
        # Special handling for live admin requests
        if any(word in message for word in ['live admin', 'human', 'speak to someone', 'real person']):
            response = "I'm connecting you with a live administrator. Please hold on...\n\n🔄 **Transferring to live support**\n\nA human representative will be with you shortly. In the meantime, you can also:\n• Call us at (043) 286-2531\n• Email info@maeslaboratory.com\n• Visit our facility during business hours"
            suggestions = ['Call now', 'Send email', 'Book appointment', 'View services', 'Check hours']
        
        # Default response if no match found
        if not response:
            if '?' in message:
                response = 'I\'d be happy to help! For specific questions not covered in my knowledge base, please:\n• Call us at (043) 286-2531\n• Email info@maeslaboratory.com\n• Request live admin support\n\nOur staff can provide detailed assistance with your inquiry.'
                suggestions = ['Request live admin', 'Call now', 'Send email', 'View services', 'Book appointment']
            else:
                response = 'Thank you for your message. I can help you with information about:\n• Laboratory services\n• Appointment booking\n• Test results\n• Payment options\n• Operating hours\n\nWhat would you like to know?'
                suggestions = ['View services', 'Book appointment', 'Check results', 'Payment info', 'Contact us']
        
        # Default suggestions if none set
        if not suggestions:
            suggestions = [
                'Book appointment',
                'View services', 
                'Check hours',
                'Payment options',
                'Contact us'
            ]
        
        # Save conversation to database
        try:
            ChatbotConversation.objects.create(
                user=request.user if request.user.is_authenticated else None,
                session_id=session_id,
                message=message,
                response=response
            )
        except Exception as e:
            print(f"Error saving chatbot conversation: {e}")
        
        return JsonResponse({
            'response': response,
            'suggestions': suggestions,
            'timestamp': timezone.now().isoformat()
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'error': 'Something went wrong. Please try again.'})

def logout_view(request):
    """Enhanced logout with audit logging"""
    if request.user.is_authenticated:
        user_name = request.user.get_full_name() or request.user.username
        create_audit_log(request, 'logout', 'User', request.user.id)
        logout(request)
        messages.success(request, f'Goodbye {user_name}! You have been logged out successfully.')
    
    return redirect('home')

# Chart generation functions
def generate_patient_appointment_chart(user):
    """Generate appointment history chart for patient"""
    try:
        # Get last 6 months of appointments
        appointments = Appointment.objects.filter(
            patient=user,
            appointment_date__gte=timezone.now() - timedelta(days=180)
        ).extra(
            select={'month': "strftime('%%Y-%%m', appointment_date)"}
        ).values('month').annotate(count=Count('id')).order_by('month')
        
        if not appointments:
            return None
        
        # Create chart
        plt.figure(figsize=(10, 6))
        months = [item['month'] for item in appointments]
        counts = [item['count'] for item in appointments]
        
        plt.bar(months, counts, color='#667eea', alpha=0.8)
        plt.title('Your Appointment History (Last 6 Months)')
        plt.xlabel('Month')
        plt.ylabel('Number of Appointments')
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        # Convert to base64
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()
        plt.close()
        
        graphic = base64.b64encode(image_png)
        graphic = graphic.decode('utf-8')
        
        return graphic
        
    except Exception as e:
        print(f"Error generating patient chart: {e}")
        return None

def generate_monthly_appointments_chart():
    """Generate monthly appointments chart for admin"""
    try:
        # Get last 12 months of data
        today = timezone.now().date()
        months_data = []
        
        for i in range(12):
            month_start = (today.replace(day=1) - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            count = Appointment.objects.filter(
                appointment_date__date__range=[month_start, month_end]
            ).count()
            months_data.append({
                'month': month_start.strftime('%b %Y'),
                'count': count
            })
        
        months_data.reverse()
        
        # Create chart
        plt.figure(figsize=(12, 6))
        months = [item['month'] for item in months_data]
        counts = [item['count'] for item in months_data]
        
        plt.plot(months, counts, marker='o', linewidth=2, markersize=6, color='#667eea')
        plt.fill_between(months, counts, alpha=0.3, color='#667eea')
        plt.title('Monthly Appointments Trend')
        plt.xlabel('Month')
        plt.ylabel('Number of Appointments')
        plt.xticks(rotation=45)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        
        # Convert to base64
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()
        plt.close()
        
        graphic = base64.b64encode(image_png)
        return graphic.decode('utf-8')
        
    except Exception as e:
        print(f"Error generating monthly appointments chart: {e}")
        return None

def generate_revenue_chart():
    """Generate revenue chart for admin"""
    try:
        # Get last 6 months of revenue data
        today = timezone.now().date()
        revenue_data = []
        
        for i in range(6):
            month_start = (today.replace(day=1) - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            revenue = Payment.objects.filter(
                payment_date__date__range=[month_start, month_end],
                payment_status='completed'
            ).aggregate(total=Sum('amount'))['total'] or 0
            revenue_data.append({
                'month': month_start.strftime('%b %Y'),
                'revenue': float(revenue)
            })
        
        revenue_data.reverse()
        
        # Create chart
        plt.figure(figsize=(10, 6))
        months = [item['month'] for item in revenue_data]
        revenues = [item['revenue'] for item in revenue_data]
        
        bars = plt.bar(months, revenues, color='#4facfe', alpha=0.8)
        plt.title('Monthly Revenue')
        plt.xlabel('Month')
        plt.ylabel('Revenue (₱)')
        plt.xticks(rotation=45)
        
        # Add value labels on bars
        for bar, revenue in zip(bars, revenues):
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                    f'₱{revenue:,.0f}', ha='center', va='bottom')
        
        plt.tight_layout()
        
        # Convert to base64
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()
        plt.close()
        
        graphic = base64.b64encode(image_png)
        return graphic.decode('utf-8')
        
    except Exception as e:
        print(f"Error generating revenue chart: {e}")
        return None

def generate_service_popularity_chart():
    """Generate service popularity chart"""
    try:
        # Get top 10 most booked services
        services = Service.objects.annotate(
            booking_count=Count('appointments')
        ).order_by('-booking_count')[:10]
        
        if not services:
            return None
        
        # Create chart
        plt.figure(figsize=(12, 8))
        service_names = [service.name[:20] + '...' if len(service.name) > 20 else service.name for service in services]
        booking_counts = [service.booking_count for service in services]
        
        bars = plt.barh(service_names, booking_counts, color='#f093fb')
        plt.title('Most Popular Services')
        plt.xlabel('Number of Bookings')
        plt.ylabel('Services')
        
        # Add value labels
        for bar, count in zip(bars, booking_counts):
            width = bar.get_width()
            plt.text(width, bar.get_y() + bar.get_height()/2.,
                    f'{count}', ha='left', va='center', fontweight='bold')
        
        plt.tight_layout()
        
        # Convert to base64
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()
        plt.close()
        
        graphic = base64.b64encode(image_png)
        return graphic.decode('utf-8')
        
    except Exception as e:
        print(f"Error generating service popularity chart: {e}")
        return None

def generate_patient_demographics_chart():
    """Generate patient demographics chart"""
    try:
        # Get age distribution
        from django.db.models import Case, When, IntegerField
        from datetime import date
        
        today = date.today()
        age_groups = UserProfile.objects.filter(
            role='patient',
            date_of_birth__isnull=False
        ).extra(
            select={
                'age': f"(julianday('now') - julianday(date_of_birth)) / 365.25"
            }
        ).extra(
            select={
                'age_group': """
                CASE 
                    WHEN age &lt; 18 THEN 'Under 18'
                    WHEN age &lt; 30 THEN '18-29'
                    WHEN age &lt; 45 THEN '30-44'
                    WHEN age &lt; 60 THEN '45-59'
                    ELSE '60+'
                END
                """
            }
        ).values('age_group').annotate(count=Count('id')).order_by('age_group')
        
        if not age_groups:
            return None
        
        # Create pie chart
        plt.figure(figsize=(8, 8))
        labels = [item['age_group'] for item in age_groups]
        sizes = [item['count'] for item in age_groups]
        colors = ['#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe']
        
        plt.pie(sizes, labels=labels, colors=colors[:len(labels)], autopct='%1.1f%%', startangle=90)
        plt.title('Patient Age Distribution')
        plt.axis('equal')
        
        # Convert to base64
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()
        plt.close()
        
        graphic = base64.b64encode(image_png)
        return graphic.decode('utf-8')
        
    except Exception as e:
        print(f"Error generating demographics chart: {e}")
        return None

# Export functions
@login_required
def export_appointments_csv(request):
    """Export appointments to CSV"""
    if request.user.userprofile.role != 'admin':
        return JsonResponse({'error': 'Access denied'})
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="appointments.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Appointment ID', 'Patient Name', 'Email', 'Phone', 'Service', 
        'Department', 'Date', 'Status', 'Payment Status', 'Amount', 
        'Discount', 'Final Amount', 'Created At'
    ])
    
    appointments = Appointment.objects.select_related(
        'patient', 'patient__userprofile', 'service', 'service__department'
    ).all()
    
    for appointment in appointments:
        writer.writerow([
            appointment.appointment_id,
            appointment.patient.get_full_name(),
            appointment.patient.email,
            appointment.patient.userprofile.phone_number,
            appointment.service.name,
            appointment.service.department.name,
            appointment.appointment_date.strftime('%Y-%m-%d %H:%M'),
            appointment.get_status_display(),
            appointment.get_payment_status_display(),
            appointment.total_amount,
            appointment.discount_amount,
            appointment.final_amount,
            appointment.created_at.strftime('%Y-%m-%d %H:%M')
        ])
    
    create_audit_log(request, 'export', 'Appointments')
    return response

@login_required
def export_appointments_excel(request):
    """Export appointments to Excel"""
    if request.user.userprofile.role != 'admin':
        return JsonResponse({'error': 'Access denied'})
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"
    
    # Headers
    headers = [
        'Appointment ID', 'Patient Name', 'Email', 'Phone', 'Service', 
        'Department', 'Date', 'Status', 'Payment Status', 'Amount', 
        'Discount', 'Final Amount', 'Created At'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    appointments = Appointment.objects.select_related(
        'patient', 'patient__userprofile', 'service', 'service__department'
    ).all()
    
    for row, appointment in enumerate(appointments, 2):
        ws.cell(row=row, column=1, value=appointment.appointment_id)
        ws.cell(row=row, column=2, value=appointment.patient.get_full_name())
        ws.cell(row=row, column=3, value=appointment.patient.email)
        ws.cell(row=row, column=4, value=appointment.patient.userprofile.phone_number)
        ws.cell(row=row, column=5, value=appointment.service.name)
        ws.cell(row=row, column=6, value=appointment.service.department.name)
        ws.cell(row=row, column=7, value=appointment.appointment_date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=8, value=appointment.get_status_display())
        ws.cell(row=row, column=9, value=appointment.get_payment_status_display())
        ws.cell(row=row, column=10, value=float(appointment.total_amount))
        ws.cell(row=row, column=11, value=float(appointment.discount_amount))
        ws.cell(row=row, column=12, value=float(appointment.final_amount))
        ws.cell(row=row, column=13, value=appointment.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="appointments.xlsx"'
    
    wb.save(response)
    create_audit_log(request, 'export', 'Appointments')
    return response

# API Views
@login_required
def get_services_by_department(request, department_id):
    """Get services filtered by department"""
    services = Service.objects.filter(
        department_id=department_id,
        is_available=True
    ).values('id', 'name', 'price', 'duration_minutes', 'requires_fasting')
    
    return JsonResponse({'services': list(services)})

@login_required
def check_appointment_availability(request):
    """Check if appointment slot is available"""
    date = request.GET.get('date')
    time = request.GET.get('time')
    service_id = request.GET.get('service_id')
    
    if not all([date, time, service_id]):
        return JsonResponse({'available': False, 'message': 'Missing parameters'})
    
    try:
        appointment_datetime = datetime.strptime(f"{date} {time}", "%Y-%m-%d %H:%M")
        appointment_datetime = timezone.make_aware(appointment_datetime)
        
        # Check if slot is taken
        existing = Appointment.objects.filter(
            appointment_date=appointment_datetime,
            service_id=service_id,
            status__in=['pending', 'confirmed']
        ).exists()
        
        return JsonResponse({
            'available': not existing,
            'message': 'Slot available' if not existing else 'Slot already taken'
        })
        
    except ValueError:
        return JsonResponse({'available': False, 'message': 'Invalid date/time format'})

# Password reset views
def password_reset_request(request):
    """Handle password reset requests"""
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            
            # Generate token
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            
            # Send email
            reset_url = request.build_absolute_uri(f'/password-reset-confirm/{uid}/{token}/')
            
            email_subject = 'Password Reset - MAES Laboratory'
            email_body = render_to_string('hospital/password_reset_email.html', {
                'user': user,
                'reset_url': reset_url,
            })
            
            send_mail(
                email_subject,
                email_body,
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )
            
            messages.success(request, 'Password reset email sent successfully!')
            
        except User.DoesNotExist:
            messages.error(request, 'No user found with this email address.')
        except Exception as e:
            messages.error(request, 'Failed to send password reset email.')
    
    return render(request, 'hospital/password_reset_form.html')

# Error handlers
def handler404(request, exception):
    """Custom 404 error page"""
    return render(request, 'hospital/404.html', status=404)

def handler500(request):
    """Custom 500 error page"""
    return render(request, 'hospital/500.html', status=500)
